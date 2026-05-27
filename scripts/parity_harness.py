#!/usr/bin/env python3
from __future__ import annotations

import json
import math
import os
import shutil
import signal
import subprocess
import sys
import time
import urllib.error
import urllib.request
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

ABS_TOL = 1e-4
REL_TOL = 1e-2
TOKEN = "regression-token"
PORT = 9002
BASE_URL = f"http://127.0.0.1:{PORT}"

ROOT = Path(__file__).resolve().parents[3]
SUITE_ROOT = ROOT / "apps" / "data-cos-suite"
RUST_ROOT = SUITE_ROOT / "rust"
TMP_ROOT = SUITE_ROOT / ".tmp" / "parity"
REPORT_PATH = SUITE_ROOT / "docs" / "parity-report.md"


def request_json(path: str, payload: dict[str, Any] | None = None) -> tuple[int, Any]:
    data = None if payload is None else json.dumps(payload).encode("utf-8")
    headers = {
        "Authorization": f"Bearer {TOKEN}",
        "Content-Type": "application/json",
    }
    req = urllib.request.Request(
        f"{BASE_URL}{path}",
        method="GET" if payload is None else "POST",
        data=data,
        headers=headers,
    )
    with urllib.request.urlopen(req, timeout=10) as resp:
        body = json.loads(resp.read().decode("utf-8") or "{}")
        return resp.status, body


def wait_health(timeout_seconds: int = 40) -> None:
    start = time.time()
    while time.time() - start < timeout_seconds:
        try:
            req = urllib.request.Request(f"{BASE_URL}/health", method="GET")
            with urllib.request.urlopen(req, timeout=2) as resp:
                if resp.status == 200:
                    return
        except urllib.error.URLError:
            time.sleep(0.2)
    raise RuntimeError("API server did not become healthy in time")


def start_server() -> subprocess.Popen[str]:
    env = os.environ.copy()
    env["DATA_COS_API_TOKEN"] = TOKEN
    env["DATA_COS_API_PORT"] = str(PORT)
    return subprocess.Popen(
        ["cargo", "run", "-p", "data-cos-api"],
        cwd=RUST_ROOT,
        env=env,
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        text=True,
    )


def stop_server(process: subprocess.Popen[str]) -> None:
    if process.poll() is not None:
        return
    process.send_signal(signal.SIGINT)
    try:
        process.wait(timeout=10)
    except subprocess.TimeoutExpired:
        process.kill()


def create_lvi_fixture(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    for idx in range(1, 19):
        ws.cell(row=idx, column=1, value=f"meta-{idx}")

    data_rows = [
        (1.0, 3.5, 1.8, 41.0),
        (2.0, 5.0, 2.0, 48.0),
        (4.0, 7.3, 2.3, 56.5),
        (0.0, 0.0, 0.0, 0.0),
    ]

    for row_offset, row_values in enumerate(data_rows, start=19):
        for col, value in enumerate(row_values, start=1):
            ws.cell(row=row_offset, column=col, value=value)

    wb.save(path)


def create_rth_fixture(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    for idx in range(1, 9):
        ws.cell(row=idx, column=1, value=f"meta-{idx}")

    # lambda, heat, current
    data_rows = [
        (964.8, 1.1, 1.0),
        (965.2, 2.2, 2.0),
        (965.7, 3.1, 4.0),
    ]

    for row_offset, row_values in enumerate(data_rows, start=9):
        for col, value in enumerate(row_values, start=1):
            ws.cell(row=row_offset, column=col, value=value)

    wb.save(path)


def create_cos_fixture(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    headers = ["LOT | SN", "仓库", "是否隔离", "ItemNum", "盒号", "货主", "冷波长", "中心波长", "2A波长"]
    ws.append(headers)

    rows = [
        ("SN-001", "良品仓", "否", "A", "B1", "模块事业部", 964.4, 964.8, 965.1),
        ("SN-002", "良品仓", "否", "A", "B1", "长光华芯", 964.5, 964.9, 965.2),
        ("SN-003", "研发工程仓", "否", "B", "B2", "模块事业部", 964.6, 965.0, 965.3),
        ("SN-004", "生产验证仓", "否", "B", "B3", "其他", 964.7, 965.1, 965.4),
        ("SN-005", "其他仓", "否", "C", "B4", "其他", 964.8, 965.2, 965.5),
        ("SN-006", "良品仓", "是", "A", "B1", "模块事业部", 964.2, 964.6, 964.9),
    ]

    for row in rows:
        ws.append(list(row))

    wb.save(path)


def create_fixture() -> dict[str, str]:
    if TMP_ROOT.exists():
        shutil.rmtree(TMP_ROOT)
    TMP_ROOT.mkdir(parents=True, exist_ok=True)

    shell_dir = TMP_ROOT / "shell_a"
    test_dir = shell_dir / "耦合测试" / "测试"
    test_dir.mkdir(parents=True, exist_ok=True)

    lvi_path = test_dir / "202501011200=LVI.xlsx"
    rth_path = test_dir / "202501011200=Rth.xlsx"

    create_lvi_fixture(lvi_path)
    create_rth_fixture(rth_path)

    cos_path = TMP_ROOT / "cos_fixture.xlsx"
    create_cos_fixture(cos_path)

    return {
        "shell": str(shell_dir),
        "cos": str(cos_path),
    }


def _read_rows(path: Path, start_row: int) -> list[list[Any]]:
    wb = load_workbook(path, data_only=True)
    ws = wb.worksheets[0]
    rows: list[list[Any]] = []
    for row in ws.iter_rows(min_row=start_row, values_only=True):
        rows.append(list(row))
    return rows


def _pick_currents(currents: list[float], requested: list[float] | None) -> set[str]:
    if requested is None:
        return {f"{v:.6f}" for v in currents}
    if len(requested) == 0:
        return {f"{max(currents):.6f}"}
    picked: set[str] = set()
    for current in currents:
        if any(abs(current - target) <= 1e-6 for target in requested):
            picked.add(f"{current:.6f}")
    return picked


def baseline_data_fetch(shell_path: str) -> list[dict[str, Any]]:
    lvi_file = Path(shell_path) / "耦合测试" / "测试" / "202501011200=LVI.xlsx"
    rth_file = Path(shell_path) / "耦合测试" / "测试" / "202501011200=Rth.xlsx"

    current_points = [2.0, 4.0]

    lvi_rows = []
    for row in _read_rows(lvi_file, 19):
        current = float(row[0]) if row[0] is not None else None
        if current is None or abs(current) <= 1e-6:
            continue
        lvi_rows.append(
            {
                "current_a": current,
                "power_w": float(row[1]) if row[1] is not None else None,
                "voltage_v": float(row[2]) if row[2] is not None else None,
                "efficiency_pct": float(row[3]) if row[3] is not None else None,
            }
        )

    selected = _pick_currents([row["current_a"] for row in lvi_rows], current_points)

    lvi_payload = [
        {
            "entry_id": shell_path,
            "test_category": "耦合测试",
            "current_a": row["current_a"],
            "power_w": row["power_w"],
            "voltage_v": row["voltage_v"],
            "efficiency_pct": row["efficiency_pct"],
            "lambda_nm": None,
            "shift_nm": None,
            "wavelength_2a_nm": None,
            "wavelength_cold_nm": None,
        }
        for row in lvi_rows
        if f"{row['current_a']:.6f}" in selected
    ]

    rth_rows = []
    for row in _read_rows(rth_file, 9):
        if row[0] is None or row[2] is None:
            continue
        current = float(row[2])
        if abs(current) <= 1e-6:
            continue
        rth_rows.append({"lambda_nm": float(row[0]), "current_a": current})

    baseline = next((row["lambda_nm"] for row in rth_rows if abs(row["current_a"] - 2.0) <= 1e-6), None)
    if baseline is None:
        baseline = min(rth_rows, key=lambda item: item["current_a"])["lambda_nm"]

    lam_2a = next((row["lambda_nm"] for row in rth_rows if abs(row["current_a"] - 2.0) <= 1e-6), None)

    x = [row["current_a"] for row in rth_rows]
    y = [row["lambda_nm"] for row in rth_rows]
    n = float(len(x))
    sum_x = sum(x)
    sum_y = sum(y)
    sum_xy = sum(a * b for a, b in zip(x, y))
    sum_x2 = sum(v * v for v in x)
    denominator = n * sum_x2 - sum_x * sum_x
    cold = None
    if abs(denominator) > 1e-12:
        cold = (sum_y * sum_x2 - sum_x * sum_xy) / denominator

    selected_rth = _pick_currents([row["current_a"] for row in rth_rows], current_points)

    rth_payload = [
        {
            "entry_id": shell_path,
            "test_category": "耦合测试",
            "current_a": row["current_a"],
            "power_w": None,
            "voltage_v": None,
            "efficiency_pct": None,
            "lambda_nm": row["lambda_nm"],
            "shift_nm": row["lambda_nm"] - baseline,
            "wavelength_2a_nm": lam_2a,
            "wavelength_cold_nm": cold,
        }
        for row in rth_rows
        if f"{row['current_a']:.6f}" in selected_rth
    ]

    merged: dict[tuple[str, str, str], dict[str, Any]] = {}
    for row in lvi_payload + rth_payload:
        key = (row["entry_id"], row["test_category"], f"{row['current_a']:.6f}")
        if key not in merged:
            merged[key] = row.copy()
            continue
        for field in [
            "power_w",
            "voltage_v",
            "efficiency_pct",
            "lambda_nm",
            "shift_nm",
            "wavelength_2a_nm",
            "wavelength_cold_nm",
        ]:
            if merged[key][field] is None and row[field] is not None:
                merged[key][field] = row[field]

    values = list(merged.values())
    values.sort(key=lambda item: item["current_a"])
    for row in values:
        for key in ["current_a", "power_w", "voltage_v", "lambda_nm", "shift_nm"]:
            if row.get(key) is not None:
                row[key] = round(float(row[key]), 3)
        if row.get("efficiency_pct") is not None:
            row["efficiency_pct"] = round(float(row["efficiency_pct"]) * 100.0, 3)
    return values


def baseline_cos_load(path: str) -> list[dict[str, Any]]:
    wb = load_workbook(path, data_only=True)
    ws = wb.worksheets[0]
    headers = [str(v).strip() for v in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    index = {name: i for i, name in enumerate(headers)}

    rows: list[dict[str, Any]] = []
    for raw in ws.iter_rows(min_row=2, values_only=True):
        device = raw[index["LOT | SN"]]
        if not device:
            continue
        rows.append(
            {
                "device_id": str(device),
                "warehouse": raw[index["仓库"]],
                "isolation": raw[index["是否隔离"]],
                "item_num": raw[index["ItemNum"]],
                "box_num": raw[index["盒号"]],
                "owner": raw[index["货主"]],
                "cold_wavelength_nm": to_float(raw[index["冷波长"]]),
                "center_wavelength_nm": to_float(raw[index["中心波长"]]),
                "two_a_wavelength_nm": to_float(raw[index["2A波长"]]),
            }
        )
    return rows


def to_float(value: Any) -> float | None:
    if value is None:
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def field_value(record: dict[str, Any], wavelength_field: str) -> float | None:
    mapping = {
        "cold": "cold_wavelength_nm",
        "center": "center_wavelength_nm",
        "two_a": "two_a_wavelength_nm",
    }
    value = record.get(mapping[wavelength_field])
    if value is None:
        return None
    return float(value)


def baseline_cos_step1(records: list[dict[str, Any]], field: str, wl_min: float, wl_max: float) -> list[dict[str, Any]]:
    usable = {"良品仓", "研发工程仓", "生产验证仓", "报废1仓"}
    low, high = sorted((wl_min, wl_max))
    out = []
    for record in records:
        if record.get("isolation") != "否":
            continue
        if record.get("warehouse") not in usable:
            continue
        wl = field_value(record, field)
        if wl is None or wl < low or wl > high:
            continue
        out.append(record)
    return out


def _pick_by_box(pool: list[dict[str, Any]], required_count: int) -> list[dict[str, Any]]:
    if required_count <= 0 or not pool:
        return []
    if len(pool) <= required_count:
        return list(pool)

    pool = sorted(pool, key=lambda row: row["_distance"])[: required_count * 2]
    boxes: dict[str, list[dict[str, Any]]] = {}
    for row in pool:
        boxes.setdefault(str(row.get("box_num") or "__missing__"), []).append(row)

    selected: list[dict[str, Any]] = []
    for _, box_rows in sorted(boxes.items(), key=lambda item: (-len(item[1]), item[0])):
        if len(selected) >= required_count:
            break
        box_rows = sorted(box_rows, key=lambda row: row["_distance"])
        selected.extend(box_rows[: required_count - len(selected)])
    return selected


def baseline_cos_step2(
    records: list[dict[str, Any]],
    field: str,
    wl_min: float,
    wl_max: float,
    required_count: int,
) -> list[dict[str, Any]]:
    target = (wl_min + wl_max) / 2.0
    scored = []
    for record in records:
        wl = field_value(record, field)
        if wl is None:
            continue
        row = dict(record)
        row["_distance"] = abs(wl - target)
        scored.append(row)

    available = [row for row in scored if row.get("warehouse") == "良品仓"]
    need_confirm = [
        row
        for row in scored
        if row.get("warehouse") in {"研发工程仓", "生产验证仓", "报废1仓"}
    ]

    selected = _pick_by_box(available, required_count)
    if len(selected) < required_count:
        selected.extend(_pick_by_box(need_confirm, required_count - len(selected)))

    for row in selected:
        row.pop("_distance", None)
    return selected


def baseline_group_greedy(
    records: list[dict[str, Any]],
    field: str,
    group_size: int,
    max_diff: float,
) -> tuple[list[list[dict[str, Any]]], list[dict[str, Any]]]:
    scored = []
    for record in records:
        wl = field_value(record, field)
        if wl is not None:
            scored.append((record, wl))
    scored.sort(key=lambda item: item[1])

    groups_idx: list[list[int]] = []
    i = 0
    while i + group_size <= len(scored):
        window = scored[i : i + group_size]
        if window[-1][1] - window[0][1] <= max_diff:
            groups_idx.append(list(range(i, i + group_size)))
            i += group_size
        else:
            i += 1

    used = {idx for group in groups_idx for idx in group}
    groups = [[scored[idx][0] for idx in group] for group in groups_idx]
    remaining = [record for idx, (record, _) in enumerate(scored) if idx not in used]
    return groups, remaining


def compare_number(path: str, left: float, right: float, errors: list[str]) -> None:
    if not (math.isfinite(left) and math.isfinite(right)):
        if left != right:
            errors.append(f"{path}: non-finite mismatch {left} vs {right}")
        return
    abs_diff = abs(left - right)
    rel_diff = abs_diff / max(abs(left), abs(right), 1.0)
    if abs_diff > ABS_TOL and rel_diff > REL_TOL:
        errors.append(f"{path}: abs={abs_diff} rel={rel_diff} ({left} vs {right})")


def compare_payload(path: str, left: Any, right: Any, errors: list[str]) -> None:
    if isinstance(left, (int, float)) and isinstance(right, (int, float)):
        compare_number(path, float(left), float(right), errors)
        return

    if isinstance(left, dict) and isinstance(right, dict):
        lk = set(left.keys())
        rk = set(right.keys())
        if lk != rk:
            errors.append(f"{path}: key mismatch {sorted(lk)} vs {sorted(rk)}")
            return
        for key in sorted(lk):
            child = f"{path}.{key}" if path else key
            compare_payload(child, left[key], right[key], errors)
        return

    if isinstance(left, list) and isinstance(right, list):
        if len(left) != len(right):
            errors.append(f"{path}: len mismatch {len(left)} vs {len(right)}")
            return
        for idx, (lv, rv) in enumerate(zip(left, right)):
            compare_payload(f"{path}[{idx}]", lv, rv, errors)
        return

    if left != right:
        errors.append(f"{path}: value mismatch {left!r} vs {right!r}")


def normalize_rows(rows: list[dict[str, Any]], key_fields: list[str]) -> list[dict[str, Any]]:
    return sorted(rows, key=lambda row: tuple(row.get(field) for field in key_fields))


def write_report(success: bool, details: list[str]) -> None:
    lines = [
        "# Parity Report",
        "",
        f"- Date: {time.strftime('%Y-%m-%d %H:%M:%S')}",
        f"- Result: {'PASS' if success else 'FAIL'}",
        f"- ABS_TOL: {ABS_TOL}",
        f"- REL_TOL: {REL_TOL}",
        "",
        "## Checks",
    ]
    lines.extend([f"- {item}" for item in details])
    REPORT_PATH.parent.mkdir(parents=True, exist_ok=True)
    REPORT_PATH.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> int:
    fixture = create_fixture()
    process = start_server()
    details: list[str] = []

    try:
        wait_health()

        # Data fetch parity
        status, rust_fetch = request_json(
            "/api/v1/data-fetch/extract",
            {
                "mode": "module",
                "entries": [fixture["shell"]],
                "test_categories": ["耦合测试"],
                "measurements": ["LVI", "Rth"],
                "current_points": [2.0, 4.0],
            },
        )
        if status != 200:
            raise RuntimeError(f"data-fetch status {status}")

        baseline_fetch = baseline_data_fetch(fixture["shell"])
        rust_rows = normalize_rows(rust_fetch["records"], ["entry_id", "test_category", "current_a"])
        py_rows = normalize_rows(baseline_fetch, ["entry_id", "test_category", "current_a"])

        errors: list[str] = []
        compare_payload("data_fetch.records", py_rows, rust_rows, errors)
        if errors:
            write_report(False, ["data-fetch parity failed"] + errors)
            for line in errors[:20]:
                print(line)
            return 1
        details.append(f"data-fetch parity ok ({len(rust_rows)} rows)")

        # COS parity
        # NOTE: /cos-filter/load stores records in process cache and only returns { total }.
        # Parity is verified by comparing step1/step2 outputs against Python baseline.
        status, rust_load = request_json("/api/v1/cos-filter/load", {"file_path": fixture["cos"]})
        if status != 200:
            raise RuntimeError(f"cos load status {status}")

        baseline_loaded = baseline_cos_load(fixture["cos"])
        py_loaded = normalize_rows(baseline_loaded, ["device_id"])

        if rust_load["total"] != len(py_loaded):
            msg = f"cos load total mismatch: rust={rust_load['total']} py={len(py_loaded)}"
            write_report(False, [msg])
            print(msg)
            return 1
        details.append(f"cos load total ok ({rust_load['total']} records)")

        # NOTE: step1 reads from process cache; request body only contains params.
        status, rust_step1 = request_json(
            "/api/v1/cos-filter/step1",
            {
                "params": {
                    "wavelength_field": "two_a",
                    "wavelength_min_nm": 965.0,
                    "wavelength_max_nm": 965.4,
                },
            },
        )
        if status != 200:
            raise RuntimeError(f"cos step1 status {status}")

        py_step1 = baseline_cos_step1(py_loaded, "two_a", 965.0, 965.4)
        rust_step1_rows = normalize_rows(rust_step1["records"], ["device_id"])
        py_step1_rows = normalize_rows(py_step1, ["device_id"])

        errors = []
        compare_payload("cos.step1", py_step1_rows, rust_step1_rows, errors)
        if errors:
            write_report(False, ["cos step1 parity failed"] + errors)
            for line in errors[:20]:
                print(line)
            return 1
        details.append(f"cos step1 parity ok ({len(rust_step1_rows)} rows)")

        status, rust_step2 = request_json(
            "/api/v1/cos-filter/step2",
            {
                "records": rust_step1_rows,
                "params": {
                    "wavelength_field": "two_a",
                    "wavelength_min_nm": 965.0,
                    "wavelength_max_nm": 965.4,
                    "required_count": 3,
                    "item_num_filter": [],
                    "box_num_filter": [],
                },
            },
        )
        if status != 200:
            raise RuntimeError(f"cos step2 status {status}")

        py_step2 = baseline_cos_step2(py_step1_rows, "two_a", 965.0, 965.4, 3)
        rust_step2_rows = normalize_rows(rust_step2["records"], ["device_id"])
        py_step2_rows = normalize_rows(py_step2, ["device_id"])

        errors = []
        compare_payload("cos.step2", py_step2_rows, rust_step2_rows, errors)
        if errors:
            write_report(False, ["cos step2 parity failed"] + errors)
            for line in errors[:20]:
                print(line)
            return 1
        details.append(f"cos step2 parity ok ({len(rust_step2_rows)} rows)")

        status, rust_group = request_json(
            "/api/v1/cos-filter/group/greedy",
            {
                "records": rust_step2_rows,
                "params": {
                    "wavelength_field": "two_a",
                    "group_size": 2,
                    "max_diff_nm": 0.3,
                    "avg_min_nm": None,
                    "avg_max_nm": None,
                },
            },
        )
        if status != 200:
            raise RuntimeError(f"cos group status {status}")

        py_groups, py_remaining = baseline_group_greedy(py_step2_rows, "two_a", 2, 0.3)

        errors = []
        compare_payload("cos.group_count", len(py_groups), rust_group["group_count"], errors)
        compare_payload("cos.remaining_count", len(py_remaining), rust_group["remaining_count"], errors)
        if errors:
            write_report(False, ["cos group parity failed"] + errors)
            for line in errors[:20]:
                print(line)
            return 1

        details.append(
            f"cos grouping parity ok ({rust_group['group_count']} groups, {rust_group['remaining_count']} remaining)"
        )

        for endpoint, params in [
            (
                "/api/v1/cos-filter/group/optimal",
                {
                    "wavelength_field": "two_a",
                    "group_size": 2,
                    "max_diff_nm": 0.3,
                    "avg_min_nm": None,
                    "avg_max_nm": None,
                },
            ),
            (
                "/api/v1/cos-filter/group/flat-top",
                {
                    "wavelength_field": "two_a",
                    "group_size": 2,
                    "max_diff_nm": 0.3,
                    "avg_min_nm": None,
                    "avg_max_nm": None,
                    "strict_mode": True,
                },
            ),
            (
                "/api/v1/cos-filter/group/huang-meng",
                {
                    "wavelength_field": "two_a",
                    "group_size": 2,
                    "max_diff_nm": 2.0,
                    "avg_min_nm": 965.1,
                    "avg_max_nm": 965.3,
                    "low_min_nm": 965.0,
                    "low_max_nm": 965.2,
                    "high_min_nm": 965.2,
                    "high_max_nm": 965.4,
                },
            ),
        ]:
            status, body = request_json(
                endpoint,
                {
                    "records": rust_step2_rows,
                    "params": params,
                },
            )
            if status != 200:
                raise RuntimeError(f"{endpoint} status {status}")
            if "group_count" not in body or "remaining_count" not in body:
                raise RuntimeError(f"{endpoint} malformed response")
            details.append(
                f"{endpoint} smoke ok ({body['group_count']} groups, {body['remaining_count']} remaining)"
            )

        write_report(True, details)
        print("parity passed")
        return 0

    finally:
        stop_server(process)


if __name__ == "__main__":
    raise SystemExit(main())
